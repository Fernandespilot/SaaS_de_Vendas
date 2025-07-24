from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, F, Count
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.forms import UserCreationForm
from decimal import Decimal
import json
from datetime import datetime, timedelta

from .models import (
    User, Cliente, PromotorVenda, Produto, Pedido, ItemPedido, 
    HistoricoStatusPedido, Municipio, GrupoProduto
)
from .forms import (
    ClienteForm, ProdutoForm, PedidoForm, ItemPedidoForm, 
    PromotorForm, CustomUserCreationForm
)
from .utils import gerar_codigo_pedido, enviar_notificacao_status


def home(request):
    """Página inicial do sistema"""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'core/home.html')


def register(request):
    """Registro de novos usuários"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, 'Conta criada com sucesso!')
            return redirect('dashboard')
    else:
        form = CustomUserCreationForm()
    return render(request, 'registration/register.html', {'form': form})


@login_required
def dashboard(request):
    """Dashboard principal baseado no tipo de usuário"""
    user = request.user
    
    context = {
        'user': user,
        'hoje': timezone.now().date(),
    }
    
    if user.tipo_usuario == 'promotor':
        promotor = get_object_or_404(PromotorVenda, user=user)
        context.update({
            'promotor': promotor,
            'clientes_area': Cliente.objects.filter(
                municipio__in=promotor.municipios_cobertura.all()
            ).count(),
            'pedidos_mes': Pedido.objects.filter(
                promotor=promotor,
                data_pedido__month=timezone.now().month
            ).count(),
            'comissao_mes': Pedido.objects.filter(
                promotor=promotor,
                data_pedido__month=timezone.now().month,
                status='concluido'
            ).aggregate(
                total=Sum(F('itens__quantidade') * F('itens__preco_unitario'))
            )['total'] or Decimal('0.00'),
        })
        return render(request, 'core/dashboard_promotor.html', context)
    
    elif user.tipo_usuario == 'cliente':
        cliente = get_object_or_404(Cliente, user=user)
        context.update({
            'cliente': cliente,
            'pedidos_pendentes': Pedido.objects.filter(
                cliente=cliente,
                status__in=['pendente', 'aprovado_estoque', 'aprovado_vendas', 'programado']
            ).count(),
            'ultimo_pedido': Pedido.objects.filter(cliente=cliente).first(),
        })
        return render(request, 'core/dashboard_cliente.html', context)
    
    elif user.tipo_usuario == 'gerente_estoque':
        context.update({
            'pedidos_avaliar': Pedido.objects.filter(status='pendente').count(),
            'pedidos_programar': Pedido.objects.filter(status='aprovado_vendas').count(),
            'entregas_hoje': Pedido.objects.filter(
                data_entrega_programada=timezone.now().date(),
                status='programado'
            ).count(),
            'produtos_estoque_baixo': Produto.objects.filter(estoque__lt=10).count(),
        })
        return render(request, 'core/dashboard_gerente_estoque.html', context)
    
    elif user.tipo_usuario == 'gerente_vendas':
        context.update({
            'pedidos_avaliar': Pedido.objects.filter(status='aprovado_estoque').count(),
            'vendas_mes': Pedido.objects.filter(
                data_pedido__month=timezone.now().month,
                status='concluido'
            ).count(),
            'receita_mes': Pedido.objects.filter(
                data_pedido__month=timezone.now().month,
                status='concluido'
            ).aggregate(
                total=Sum(F('itens__quantidade') * F('itens__preco_unitario'))
            )['total'] or Decimal('0.00'),
        })
        return render(request, 'core/dashboard_gerente_vendas.html', context)
    
    elif user.tipo_usuario == 'gerenciador':
        context.update({
            'total_promotores': PromotorVenda.objects.count(),
            'total_clientes': Cliente.objects.count(),
            'total_produtos': Produto.objects.count(),
            'pedidos_total': Pedido.objects.count(),
        })
        return render(request, 'core/dashboard_gerenciador.html', context)
    
    # Dashboard padrão para outros tipos de usuário
    return render(request, 'core/dashboard.html', context)


@login_required
def listar_clientes(request):
    """Lista clientes para promotores"""
    if request.user.tipo_usuario != 'promotor':
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    promotor = get_object_or_404(PromotorVenda, user=request.user)
    clientes = Cliente.objects.filter(
        municipio__in=promotor.municipios_cobertura.all()
    ).order_by('user__first_name')
    
    # Filtros
    busca = request.GET.get('busca', '')
    ordenacao = request.GET.get('ordenacao', 'nome')
    
    if busca:
        clientes = clientes.filter(
            Q(user__first_name__icontains=busca) |
            Q(user__last_name__icontains=busca) |
            Q(endereco__icontains=busca)
        )
    
    if ordenacao == 'endereco':
        clientes = clientes.order_by('endereco')
    elif ordenacao == 'municipio':
        clientes = clientes.order_by('municipio__nome')
    else:
        clientes = clientes.order_by('user__first_name')
    
    paginator = Paginator(clientes, 10)
    page = request.GET.get('page')
    clientes = paginator.get_page(page)
    
    return render(request, 'core/listar_clientes.html', {
        'clientes': clientes,
        'busca': busca,
        'ordenacao': ordenacao,
    })


@login_required
def cadastrar_cliente(request):
    """Cadastro de novos clientes"""
    if request.user.tipo_usuario not in ['promotor', 'gerenciador']:
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = ClienteForm(request.POST)
        if form.is_valid():
            cliente = form.save()
            if request.user.tipo_usuario == 'promotor':
                promotor = get_object_or_404(PromotorVenda, user=request.user)
                cliente.promotor = promotor
                cliente.save()
            
            messages.success(request, 'Cliente cadastrado com sucesso!')
            return redirect('listar_clientes')
    else:
        form = ClienteForm()
    
    return render(request, 'core/cadastrar_cliente.html', {'form': form})


@login_required
def listar_produtos(request):
    """Lista produtos disponíveis"""
    produtos = Produto.objects.filter(ativo=True).order_by('nome')
    
    # Filtros
    busca = request.GET.get('busca', '')
    grupo = request.GET.get('grupo', '')
    
    if busca:
        produtos = produtos.filter(
            Q(nome__icontains=busca) |
            Q(codigo__icontains=busca)
        )
    
    if grupo:
        produtos = produtos.filter(grupo_id=grupo)
    
    grupos = GrupoProduto.objects.all()
    
    paginator = Paginator(produtos, 12)
    page = request.GET.get('page')
    produtos = paginator.get_page(page)
    
    return render(request, 'core/listar_produtos.html', {
        'produtos': produtos,
        'grupos': grupos,
        'busca': busca,
        'grupo_selecionado': grupo,
    })


@login_required
def criar_pedido(request):
    """Criação de novos pedidos"""
    if request.user.tipo_usuario != 'promotor':
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    promotor = get_object_or_404(PromotorVenda, user=request.user)
    
    if request.method == 'POST':
        form = PedidoForm(request.POST, promotor=promotor)
        if form.is_valid():
            pedido = form.save(commit=False)
            pedido.promotor = promotor
            pedido.codigo = gerar_codigo_pedido()
            pedido.save()
            
            # Processar itens do pedido
            itens_data = json.loads(request.POST.get('itens', '[]'))
            
            for item_data in itens_data:
                produto = get_object_or_404(Produto, id=item_data['produto_id'])
                ItemPedido.objects.create(
                    pedido=pedido,
                    produto=produto,
                    quantidade=item_data['quantidade'],
                    preco_unitario=produto.preco_final
                )
            
            messages.success(request, f'Pedido {pedido.codigo} criado com sucesso!')
            return redirect('listar_pedidos')
    else:
        form = PedidoForm(promotor=promotor)
    
    return render(request, 'core/criar_pedido.html', {'form': form})


@login_required
def listar_pedidos(request):
    """Lista pedidos baseado no tipo de usuário"""
    user = request.user
    
    if user.tipo_usuario == 'promotor':
        promotor = get_object_or_404(PromotorVenda, user=user)
        pedidos = Pedido.objects.filter(promotor=promotor)
    elif user.tipo_usuario == 'cliente':
        cliente = get_object_or_404(Cliente, user=user)
        pedidos = Pedido.objects.filter(cliente=cliente)
    elif user.tipo_usuario in ['gerente_estoque', 'gerente_vendas', 'gerenciador']:
        pedidos = Pedido.objects.all()
    else:
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    # Filtros
    status = request.GET.get('status', '')
    data_inicio = request.GET.get('data_inicio', '')
    data_fim = request.GET.get('data_fim', '')
    
    if status:
        pedidos = pedidos.filter(status=status)
    
    if data_inicio:
        pedidos = pedidos.filter(data_pedido__gte=data_inicio)
    
    if data_fim:
        pedidos = pedidos.filter(data_pedido__lte=data_fim)
    
    pedidos = pedidos.order_by('-data_pedido')
    
    paginator = Paginator(pedidos, 10)
    page = request.GET.get('page')
    pedidos = paginator.get_page(page)
    
    return render(request, 'core/listar_pedidos.html', {
        'pedidos': pedidos,
        'status_choices': Pedido.STATUS_CHOICES,
        'status_selecionado': status,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
    })


@login_required
def detalhar_pedido(request, pedido_id):
    """Detalhes de um pedido específico"""
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    # Verificar permissão
    if request.user.tipo_usuario == 'promotor':
        promotor = get_object_or_404(PromotorVenda, user=request.user)
        if pedido.promotor != promotor:
            messages.error(request, 'Acesso negado.')
            return redirect('listar_pedidos')
    elif request.user.tipo_usuario == 'cliente':
        cliente = get_object_or_404(Cliente, user=request.user)
        if pedido.cliente != cliente:
            messages.error(request, 'Acesso negado.')
            return redirect('listar_pedidos')
    
    historico = HistoricoStatusPedido.objects.filter(pedido=pedido).order_by('-data_mudanca')
    
    return render(request, 'core/detalhar_pedido.html', {
        'pedido': pedido,
        'historico': historico,
    })


@login_required
def avaliar_pedido_estoque(request, pedido_id):
    """Avaliação de pedidos pelo gerente de estoque"""
    if request.user.tipo_usuario != 'gerente_estoque':
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        observacoes = request.POST.get('observacoes', '')
        
        if acao == 'aprovar':
            pedido.status = 'aprovado_estoque'
            messages.success(request, 'Pedido aprovado pelo estoque!')
        elif acao == 'reprovar':
            pedido.status = 'reprovado_estoque'
            messages.success(request, 'Pedido reprovado pelo estoque!')
        
        pedido.save()
        
        # Registrar no histórico
        HistoricoStatusPedido.objects.create(
            pedido=pedido,
            status_anterior='pendente',
            status_novo=pedido.status,
            usuario=request.user,
            observacoes=observacoes
        )
        
        # Enviar notificação
        enviar_notificacao_status(pedido)
        
        return redirect('listar_pedidos')
    
    return render(request, 'core/avaliar_pedido_estoque.html', {'pedido': pedido})


@login_required
def avaliar_pedido_vendas(request, pedido_id):
    """Avaliação de pedidos pelo gerente de vendas"""
    if request.user.tipo_usuario != 'gerente_vendas':
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        acao = request.POST.get('acao')
        observacoes = request.POST.get('observacoes', '')
        
        if acao == 'aprovar':
            pedido.status = 'aprovado_vendas'
            messages.success(request, 'Pedido aprovado pelas vendas!')
        elif acao == 'reprovar':
            pedido.status = 'reprovado_vendas'
            messages.success(request, 'Pedido reprovado pelas vendas!')
        
        pedido.save()
        
        # Registrar no histórico
        HistoricoStatusPedido.objects.create(
            pedido=pedido,
            status_anterior='aprovado_estoque',
            status_novo=pedido.status,
            usuario=request.user,
            observacoes=observacoes
        )
        
        # Enviar notificação
        enviar_notificacao_status(pedido)
        
        return redirect('listar_pedidos')
    
    return render(request, 'core/avaliar_pedido_vendas.html', {'pedido': pedido})


@login_required
def programar_entrega(request, pedido_id):
    """Programação de entrega pelo gerente de estoque"""
    if request.user.tipo_usuario != 'gerente_estoque':
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        data_entrega = request.POST.get('data_entrega')
        
        if data_entrega:
            pedido.data_entrega_programada = data_entrega
            pedido.status = 'programado'
            pedido.save()
            
            # Atualizar estoque
            for item in pedido.itens.all():
                if item.produto.estoque >= item.quantidade:
                    item.produto.estoque -= item.quantidade
                    item.produto.save()
                else:
                    messages.error(request, f'Estoque insuficiente para {item.produto.nome}')
                    return redirect('programar_entrega', pedido_id=pedido_id)
            
            # Registrar no histórico
            HistoricoStatusPedido.objects.create(
                pedido=pedido,
                status_anterior='aprovado_vendas',
                status_novo='programado',
                usuario=request.user,
                observacoes=f'Entrega programada para {data_entrega}'
            )
            
            # Enviar notificação
            enviar_notificacao_status(pedido)
            
            messages.success(request, 'Entrega programada com sucesso!')
            return redirect('listar_pedidos')
    
    return render(request, 'core/programar_entrega.html', {'pedido': pedido})


@login_required
def processar_entrega(request, pedido_id):
    """Processamento de entrega pelo gerente de estoque"""
    if request.user.tipo_usuario != 'gerente_estoque':
        messages.error(request, 'Acesso negado.')
        return redirect('dashboard')
    
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    if request.method == 'POST':
        pedido.status = 'processado'
        pedido.save()
        
        # Registrar no histórico
        HistoricoStatusPedido.objects.create(
            pedido=pedido,
            status_anterior='programado',
            status_novo='processado',
            usuario=request.user,
            observacoes='Entrega processada'
        )
        
        # Enviar notificação
        enviar_notificacao_status(pedido)
        
        messages.success(request, 'Entrega processada com sucesso!')
        return redirect('listar_pedidos')
    
    return render(request, 'core/processar_entrega.html', {'pedido': pedido})


@login_required
def get_produto_info(request, produto_id):
    """API para obter informações de um produto"""
    produto = get_object_or_404(Produto, id=produto_id)
    
    data = {
        'id': produto.id,
        'nome': produto.nome,
        'codigo': produto.codigo,
        'preco': float(produto.preco_final),
        'estoque': produto.estoque,
        'grupo': produto.grupo.nome,
    }
    
    return JsonResponse(data)


@login_required
def buscar_produtos(request):
    """API para buscar produtos"""
    termo = request.GET.get('q', '')
    
    produtos = Produto.objects.filter(
        Q(nome__icontains=termo) | Q(codigo__icontains=termo),
        ativo=True
    )[:10]
    
    data = [
        {
            'id': p.id,
            'nome': p.nome,
            'codigo': p.codigo,
            'preco': float(p.preco_final),
            'estoque': p.estoque,
        }
        for p in produtos
    ]
    
    return JsonResponse(data, safe=False)

@login_required
def relatorios(request):
    """Gera relatórios baseados nos filtros"""
    dados_vendas = None
    dados_estoque = None
    dados_clientes = None
    vendas_por_promotor = None
    vendas_por_produto = None
    periodo_display = None
    grafico_vendas = None
    grafico_categoria = None
    
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    tipo_relatorio = request.GET.get('tipo_relatorio')
    formato = request.GET.get('formato', 'html')
    
    if data_inicio:
        data_inicio = datetime.strptime(data_inicio, '%Y-%m-%d').date()
    if data_fim:
        data_fim = datetime.strptime(data_fim, '%Y-%m-%d').date()
    
    # Relatório de Vendas
    if tipo_relatorio == 'vendas' or not tipo_relatorio:
        pedidos = Pedido.objects.filter(status='aprovado')
        if data_inicio:
            pedidos = pedidos.filter(data_pedido__gte=data_inicio)
        if data_fim:
            pedidos = pedidos.filter(data_pedido__lte=data_fim)
        
        dados_vendas = {
            'total_pedidos': pedidos.count(),
            'valor_total': pedidos.aggregate(total=Sum('valor_total'))['total'] or 0,
            'comissao_total': pedidos.aggregate(total=Sum('comissao_total'))['total'] or 0,
            'ticket_medio': 0
        }
        
        if dados_vendas['total_pedidos'] > 0:
            dados_vendas['ticket_medio'] = dados_vendas['valor_total'] / dados_vendas['total_pedidos']
        
        # Vendas por promotor
        vendas_por_promotor = pedidos.values('promotor__user__first_name', 'promotor__user__last_name').annotate(
            total_pedidos=Count('id'),
            valor_total=Sum('valor_total'),
            comissao_total=Sum('comissao_total')
        ).order_by('-valor_total')
        
        for promotor in vendas_por_promotor:
            promotor['nome'] = f"{promotor['promotor__user__first_name']} {promotor['promotor__user__last_name']}"
        
        # Vendas por produto
        vendas_por_produto = ItemPedido.objects.filter(pedido__in=pedidos).values('produto__nome').annotate(
            quantidade_vendida=Sum('quantidade'),
            valor_total=Sum('subtotal')
        ).order_by('-valor_total')[:10]
        
        total_vendas = vendas_por_produto.aggregate(total=Sum('valor_total'))['total'] or 0
        for produto in vendas_por_produto:
            produto['nome'] = produto['produto__nome']
            produto['percentual'] = (produto['valor_total'] / total_vendas * 100) if total_vendas > 0 else 0
        
        # Período
        if data_inicio and data_fim:
            periodo_display = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
        else:
            periodo_display = "Todos os períodos"
    
    # Relatório de Estoque
    if tipo_relatorio == 'estoque' or not tipo_relatorio:
        produtos = Produto.objects.all()
        dados_estoque = []
        
        for produto in produtos:
            dados_estoque.append({
                'nome': produto.nome,
                'codigo': produto.codigo,
                'estoque_atual': produto.estoque_atual,
                'estoque_minimo': produto.estoque_minimo,
                'valor_estoque': produto.estoque_atual * produto.preco_venda
            })
    
    # Relatório de Clientes
    if tipo_relatorio == 'clientes' or not tipo_relatorio:
        clientes = Cliente.objects.select_related('user', 'municipio').all()
        dados_clientes = []
        
        for cliente in clientes:
            pedidos_cliente = Pedido.objects.filter(cliente=cliente)
            ultimo_pedido = pedidos_cliente.order_by('-data_pedido').first()
            
            dados_clientes.append({
                'nome': cliente.user.get_full_name(),
                'municipio': cliente.municipio.nome if cliente.municipio else 'Não definido',
                'total_pedidos': pedidos_cliente.count(),
                'valor_total': pedidos_cliente.aggregate(total=Sum('valor_total'))['total'] or 0,
                'ultimo_pedido': ultimo_pedido.data_pedido if ultimo_pedido else None,
                'status_financeiro': cliente.status_financeiro
            })
    
    # Gerar PDF ou Excel se solicitado
    if formato == 'pdf':
        return gerar_relatorio_pdf(dados_vendas, dados_estoque, dados_clientes, periodo_display)
    elif formato == 'excel':
        return gerar_relatorio_excel(dados_vendas, dados_estoque, dados_clientes, periodo_display)
    
    # Dados para gráficos
    if dados_vendas:
        grafico_vendas = {
            'labels': ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun'],
            'data': [1000, 1500, 1200, 1800, 2000, 1600]
        }
        
        grafico_categoria = {
            'labels': ['Eletrônicos', 'Roupas', 'Casa', 'Livros'],
            'data': [30, 25, 25, 20]
        }
    
    context = {
        'dados_vendas': dados_vendas,
        'dados_estoque': dados_estoque,
        'dados_clientes': dados_clientes,
        'vendas_por_promotor': vendas_por_promotor,
        'vendas_por_produto': vendas_por_produto,
        'periodo_display': periodo_display,
        'grafico_vendas': grafico_vendas,
        'grafico_categoria': grafico_categoria,
    }
    
    return render(request, 'core/relatorios.html', context)

@login_required
def perfil(request):
    """Gerencia o perfil do usuário"""
    if request.method == 'POST':
        acao = request.POST.get('acao')
        
        if acao == 'atualizar_perfil':
            request.user.first_name = request.POST.get('first_name')
            request.user.last_name = request.POST.get('last_name')
            request.user.email = request.POST.get('email')
            request.user.telefone = request.POST.get('telefone')
            request.user.cpf = request.POST.get('cpf')
            request.user.endereco = request.POST.get('endereco')
            
            data_nascimento = request.POST.get('data_nascimento')
            if data_nascimento:
                request.user.data_nascimento = datetime.strptime(data_nascimento, '%Y-%m-%d').date()
            
            request.user.save()
            messages.success(request, 'Perfil atualizado com sucesso!')
            
        elif acao == 'alterar_senha':
            senha_atual = request.POST.get('senha_atual')
            nova_senha = request.POST.get('nova_senha')
            confirmar_senha = request.POST.get('confirmar_senha')
            
            if not request.user.check_password(senha_atual):
                messages.error(request, 'Senha atual incorreta!')
            elif nova_senha != confirmar_senha:
                messages.error(request, 'As senhas não coincidem!')
            else:
                request.user.set_password(nova_senha)
                request.user.save()
                messages.success(request, 'Senha alterada com sucesso!')
                return redirect('login')
        
        return redirect('perfil')
    
    # Estatísticas do usuário
    estatisticas = {}
    if request.user.tipo_usuario == 'promotor':
        pedidos_mes = Pedido.objects.filter(
            promotor__user=request.user,
            data_pedido__month=datetime.now().month
        )
        estatisticas = {
            'pedidos_mes': pedidos_mes.count(),
            'comissao_mes': pedidos_mes.aggregate(total=Sum('comissao_total'))['total'] or 0
        }
    elif request.user.tipo_usuario == 'cliente':
        pedidos = Pedido.objects.filter(cliente__user=request.user)
        estatisticas = {
            'total_pedidos': pedidos.count(),
            'valor_total': pedidos.aggregate(total=Sum('valor_total'))['total'] or 0
        }
    
    context = {
        'estatisticas': estatisticas
    }
    
    return render(request, 'core/perfil.html', context)

def listar_municipios(request):
    """Lista municípios para select dinâmico"""
    estado = request.GET.get('estado')
    municipios = Municipio.objects.filter(estado=estado).order_by('nome')
    
    data = [{'id': m.id, 'nome': m.nome} for m in municipios]
    return JsonResponse(data, safe=False)

def gerar_relatorio_pdf(dados_vendas, dados_estoque, dados_clientes, periodo):
    """Gera relatório em PDF"""
    from django.http import HttpResponse
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio.pdf"'
    
    p = canvas.Canvas(response, pagesize=letter)
    
    # Título
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, 750, "Relatório SisVenda")
    
    # Período
    p.setFont("Helvetica", 12)
    p.drawString(100, 720, f"Período: {periodo}")
    
    # Dados de vendas
    if dados_vendas:
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 680, "Vendas")
        
        y = 650
        p.setFont("Helvetica", 10)
        p.drawString(100, y, f"Total de Pedidos: {dados_vendas['total_pedidos']}")
        y -= 20
        p.drawString(100, y, f"Valor Total: R$ {dados_vendas['valor_total']:.2f}")
        y -= 20
        p.drawString(100, y, f"Ticket Médio: R$ {dados_vendas['ticket_medio']:.2f}")
    
    p.showPage()
    p.save()
    
    return response

def gerar_relatorio_excel(dados_vendas, dados_estoque, dados_clientes, periodo):
    """Gera relatório em Excel"""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.utils import get_column_letter
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    # Cabeçalho
    ws['A1'] = 'Relatório SisVenda'
    ws['A2'] = f'Período: {periodo}'
    
    # Dados de vendas
    if dados_vendas:
        row = 4
        ws[f'A{row}'] = 'VENDAS'
        row += 1
        ws[f'A{row}'] = 'Total de Pedidos'
        ws[f'B{row}'] = dados_vendas['total_pedidos']
        row += 1
        ws[f'A{row}'] = 'Valor Total'
        ws[f'B{row}'] = dados_vendas['valor_total']
        row += 1
        ws[f'A{row}'] = 'Ticket Médio'
        ws[f'B{row}'] = dados_vendas['ticket_medio']
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio.xlsx"'
    wb.save(response)
    
    return response
