"""
Views para o app de relatórios
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta
from apps.pedidos.models import Pedido, ItemPedido
from apps.produtos.models import Produto
from apps.usuarios.models import User
import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

@login_required
def relatorios_dashboard(request):
    """Dashboard principal de relatórios"""
    # Estatísticas gerais
    total_vendas = Pedido.objects.filter(status='finalizado').aggregate(
        total=Sum('valor_total')
    )['total'] or 0
    
    total_pedidos = Pedido.objects.count()
    total_produtos = Produto.objects.count()
    total_usuarios = User.objects.count()
    
    # Vendas por mês
    hoje = timezone.now().date()
    inicio_mes = hoje.replace(day=1)
    vendas_mes = Pedido.objects.filter(
        data_criacao__gte=inicio_mes,
        status='finalizado'
    ).aggregate(total=Sum('valor_total'))['total'] or 0
    
    context = {
        'total_vendas': total_vendas,
        'total_pedidos': total_pedidos,
        'total_produtos': total_produtos,
        'total_usuarios': total_usuarios,
        'vendas_mes': vendas_mes,
    }
    return render(request, 'relatorios/dashboard.html', context)

@login_required
def relatorio_vendas(request):
    """Relatório de vendas"""
    # Filtros
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')
    status = request.GET.get('status')
    
    pedidos = Pedido.objects.select_related('usuario', 'vendedor').order_by('-data_criacao')
    
    if data_inicio:
        pedidos = pedidos.filter(data_criacao__gte=data_inicio)
    if data_fim:
        pedidos = pedidos.filter(data_criacao__lte=data_fim)
    if status:
        pedidos = pedidos.filter(status=status)
    
    # Estatísticas
    total_vendas = pedidos.aggregate(total=Sum('valor_total'))['total'] or 0
    total_pedidos = pedidos.count()
    
    context = {
        'pedidos': pedidos,
        'total_vendas': total_vendas,
        'total_pedidos': total_pedidos,
        'status_choices': Pedido.STATUS_CHOICES,
    }
    return render(request, 'relatorios/vendas.html', context)

@login_required
def relatorio_produtos(request):
    """Relatório de produtos"""
    produtos = Produto.objects.annotate(
        total_vendido=Sum('itempedido__quantidade')
    ).order_by('-total_vendido')
    
    context = {
        'produtos': produtos,
    }
    return render(request, 'relatorios/produtos.html', context)

@login_required
def relatorio_usuarios(request):
    """Relatório de usuários"""
    usuarios = User.objects.annotate(
        total_pedidos=Count('pedido')
    ).order_by('-total_pedidos')
    
    context = {
        'usuarios': usuarios,
    }
    return render(request, 'relatorios/usuarios.html', context)

@login_required
def relatorio_vendas_pdf(request):
    """Exporta relatório de vendas em PDF"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relatorio_vendas.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title = Paragraph("Relatório de Vendas", styles['Title'])
    story.append(title)
    
    # Dados
    pedidos = Pedido.objects.select_related('usuario', 'vendedor').order_by('-data_criacao')
    
    data = [['ID', 'Cliente', 'Vendedor', 'Status', 'Total', 'Data']]
    for pedido in pedidos:
        data.append([
            str(pedido.id),
            pedido.usuario.username,
            pedido.vendedor.username,
            pedido.get_status_display(),
            f'R$ {pedido.valor_total:.2f}',
            pedido.data_criacao.strftime('%d/%m/%Y')
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    doc.build(story)
    
    return response

@login_required
def relatorio_vendas_excel(request):
    """Exporta relatório de vendas em Excel"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="relatorio_vendas.xlsx"'
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Relatório de Vendas'
    
    # Cabeçalhos
    headers = ['ID', 'Cliente', 'Vendedor', 'Status', 'Total', 'Data']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)
    
    # Dados
    pedidos = Pedido.objects.select_related('usuario', 'vendedor').order_by('-data_criacao')
    
    for row, pedido in enumerate(pedidos, 2):
        worksheet.cell(row=row, column=1, value=pedido.id)
        worksheet.cell(row=row, column=2, value=pedido.usuario.username)
        worksheet.cell(row=row, column=3, value=pedido.vendedor.username)
        worksheet.cell(row=row, column=4, value=pedido.get_status_display())
        worksheet.cell(row=row, column=5, value=float(pedido.valor_total))
        worksheet.cell(row=row, column=6, value=pedido.data_criacao.strftime('%d/%m/%Y'))
    
    workbook.save(response)
    return response
